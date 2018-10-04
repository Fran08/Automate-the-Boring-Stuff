#! python3
# update_produce.py - Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

# Loop through the rows and update the prices.
for row_num in range(2, sheet.max_row + 1):
    produce = sheet['A' + str(row_num)].value
    # Below is just showing another way to write it, commented out.
    # produce = sheet.cell(row=row_num, column=1).value

    if produce in PRICE_UPDATES:
        sheet['B' + str(row_num)].value = PRICE_UPDATES[produce]
        # sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce]

wb.save('UpdatedProduceSales.xlsx')
