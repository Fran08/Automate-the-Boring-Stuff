#! python3
# multiplication_table.py - Takes an integer (n) as an argument and returns a
# multiplication table of n x n size.

import sys, openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

if sys.argv[1].isdigit():
    n = int(sys.argv[1])

    wb = openpyxl.Workbook()
    sheet = wb.active
    bold_font = Font(b=True)

    for i in range(1, n+1):
        sheet['A' + str(i + 1)] = i
        sheet['A' + str(i + 1)].font = bold_font
        sheet[get_column_letter(i + 1) + '1'] = i
        sheet[get_column_letter(i + 1) + '1'].font = bold_font

    max_col = get_column_letter(sheet.max_column)

    for row_obj in sheet['B2': max_col + str(n+1)]:
        for cell in row_obj:
            cell_col = cell.column
            cell_row = cell.row

            x = 'A' + str(cell_row)
            y = cell_col + '1'
            cell.value = f'={x} * {y}'

    wb.save('multiplication_table.xlsx')
    print('File saved.')

else:
    print('You did not enter an integer value.')
