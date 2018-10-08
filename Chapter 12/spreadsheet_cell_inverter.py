#! python3
# spreadsheet_cell_inverter.py - Invert the row and column of cells in spreadsheet
# Kept it simple, you can definitely add in more features if you would like.

import openpyxl

print('Enter spreadsheet to invert:')
filename = input()

wb = openpyxl.load_workbook(filename)
sheet = wb.active
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

for x in range(1, sheet.max_row + 1):
    for y in range(1, sheet.max_column + 1):
        new_sheet.cell(row=y, column=x).value = sheet.cell(row=x, column=y).value

new_wb.save('inverter.xlsx')
print('inverter.xlsx saved.')
