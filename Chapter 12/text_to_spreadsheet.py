#! python3
# text_to_spreadsheet.py - Reads in text files and puts the contents of each
# on one row in an Excel spreadsheet

import openpyxl, os
from openpyxl.utils import get_column_letter

# This line used to create some random text files.
##for i in range(10):
##    doc = open(f'spam{str(i+1).zfill(2)}.txt', 'w')
##    doc.write(f'This is some random line of text for spam {i+1}')
##    doc.close()

wb = openpyxl.Workbook()
sheet = wb.active

folder = os.path.abspath(input('Please enter folder to search: '))

if os.path.isdir(folder) is True:
    prefix = input('Enter file prefix: ')
    files = os.listdir(folder)
    row_num = 1
    
    for file in files:
        if file.startswith(prefix):
            lines = open(os.path.join(folder, file)).readlines()

            for line in lines:
                sheet.cell(row=row_num, column=1).value = line
                row_num += 1

    wb.save(os.path.join(folder, 'text_to_sheet.xlsx'))
