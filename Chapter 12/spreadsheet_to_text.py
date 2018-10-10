#! python3
# spreadsheet_to_text.py - Reads in an Excel worksheet and saves the cells of
# column A into one text file, cells of column B into another, and so on.

import openpyxl, os

wb = openpyxl.Workbook()
sheet = wb.active

spreadsheet = os.path.abspath(input('Please enter spreadsheet to split: '))

if os.path.isfile(spreadsheet) is True:
    path = os.path.dirname(spreadsheet)
    prefix = input('Enter text document prefix: ')
    wb = openpyxl.load_workbook(spreadsheet)
    sheet = wb.active

    col_num = 1
    
    for i in range(1, sheet.max_column + 1):
        
        col_data = []

        for r in range(1, sheet.max_row + 1):
            if sheet.cell(row = r, column=col_num).value:
                col_data.append(sheet.cell(row=r, column=col_num).value)

        file_name = f'{prefix}{col_num}.txt'
        file = open(os.path.join(path, file_name), 'w')

        for line in col_data:
            file.write(f'{line}\n')
        file.close()

        col_num += 1
